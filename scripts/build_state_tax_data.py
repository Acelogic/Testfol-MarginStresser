#!/usr/bin/env python3
"""
Build state income tax bracket data from authoritative sources.

Outputs: data/state_income_tax_rates.json

Sources:
  Phase A: TAXSIM35 (NBER) — 1977–2013 (opt-in, downloads external binary)
  Phase B: taxgraphs (GitHub) — 2014–2024
  Phase C: Tax Foundation Excel — 2023–2026

Usage:
  python scripts/build_state_tax_data.py                     # Phases B+C only
  python scripts/build_state_tax_data.py --include-taxsim    # All phases
  python scripts/build_state_tax_data.py --skip-excel        # Phase B only
"""

import argparse
import json
import logging
import os
import re
import subprocess
import urllib.request
from collections import defaultdict

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
log = logging.getLogger(__name__)

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(PROJECT_ROOT, "data")
OUTPUT_PATH = os.path.join(DATA_DIR, "state_income_tax_rates.json")
CACHE_DIR = os.path.join(PROJECT_ROOT, "scripts", ".cache")

# ─── State Constants ─────────────────────────────────────────────────────────

STATE_NAMES = {
    "AL": "Alabama", "AK": "Alaska", "AZ": "Arizona", "AR": "Arkansas",
    "CA": "California", "CO": "Colorado", "CT": "Connecticut", "DE": "Delaware",
    "DC": "District of Columbia", "FL": "Florida", "GA": "Georgia", "HI": "Hawaii",
    "ID": "Idaho", "IL": "Illinois", "IN": "Indiana", "IA": "Iowa",
    "KS": "Kansas", "KY": "Kentucky", "LA": "Louisiana", "ME": "Maine",
    "MD": "Maryland", "MA": "Massachusetts", "MI": "Michigan", "MN": "Minnesota",
    "MS": "Mississippi", "MO": "Missouri", "MT": "Montana", "NE": "Nebraska",
    "NV": "Nevada", "NH": "New Hampshire", "NJ": "New Jersey", "NM": "New Mexico",
    "NY": "New York", "NC": "North Carolina", "ND": "North Dakota", "OH": "Ohio",
    "OK": "Oklahoma", "OR": "Oregon", "PA": "Pennsylvania", "RI": "Rhode Island",
    "SC": "South Carolina", "SD": "South Dakota", "TN": "Tennessee", "TX": "Texas",
    "UT": "Utah", "VT": "Vermont", "VA": "Virginia", "WA": "Washington",
    "WV": "West Virginia", "WI": "Wisconsin", "WY": "Wyoming",
}

# Reverse lookup: lowercased name → code
NAME_TO_CODE = {name.lower(): code for code, name in STATE_NAMES.items()}

STATE_FIPS = {
    "AL": 1, "AK": 2, "AZ": 3, "AR": 4, "CA": 5, "CO": 6, "CT": 7, "DE": 8,
    "DC": 9, "FL": 10, "GA": 11, "HI": 12, "ID": 13, "IL": 14, "IN": 15,
    "IA": 16, "KS": 17, "KY": 18, "LA": 19, "ME": 20, "MD": 21, "MA": 22,
    "MI": 23, "MN": 24, "MS": 25, "MO": 26, "MT": 27, "NE": 28, "NV": 29,
    "NH": 30, "NJ": 31, "NM": 32, "NY": 33, "NC": 34, "ND": 35, "OH": 36,
    "OK": 37, "OR": 38, "PA": 39, "RI": 40, "SC": 41, "SD": 42, "TN": 43,
    "TX": 44, "UT": 45, "VT": 46, "VA": 47, "WA": 48, "WV": 49, "WI": 50,
    "WY": 51,
}
FIPS_TO_CODE = {v: k for k, v in STATE_FIPS.items()}

NO_TAX_STATES = {"AK", "FL", "NV", "NH", "SD", "TN", "TX", "WA", "WY"}

LT_EXCLUSION_PCT = {
    "AZ": 0.25, "AR": 0.50, "MT": 0.40, "NM": 0.40,
    "ND": 0.40, "SC": 0.44, "VT": 0.40, "WI": 0.30,
}

# Income probes for TAXSIM bracket detection (denser at low end)
INCOME_PROBES = [
    0, 500, 1000, 1500, 2000, 2500, 3000, 3500, 4000, 4500, 5000,
    6000, 7000, 8000, 9000, 10000, 12000, 14000, 16000, 18000, 20000,
    22000, 24000, 26000, 28000, 30000, 35000, 40000, 45000, 50000,
    55000, 60000, 65000, 70000, 75000, 80000, 90000, 100000, 110000,
    120000, 130000, 140000, 150000, 175000, 200000, 250000, 300000,
    400000, 500000, 600000, 750000, 1000000, 1500000, 2000000,
    5000000, 10000000,
]


STATE_ABBREVS = {
    "ala": "AL", "alaska": "AK", "ariz": "AZ", "ark": "AR", "calif": "CA",
    "colo": "CO", "conn": "CT", "del": "DE", "d.c": "DC", "fla": "FL",
    "ga": "GA", "hawaii": "HI", "idaho": "ID", "ill": "IL", "ind": "IN",
    "iowa": "IA", "kans": "KS", "ky": "KY", "la": "LA", "maine": "ME",
    "md": "MD", "mass": "MA", "mich": "MI", "minn": "MN", "miss": "MS",
    "mo": "MO", "mont": "MT", "nebr": "NE", "neb": "NE", "nev": "NV",
    "n.h": "NH", "n.j": "NJ", "n.m": "NM", "n.y": "NY", "n.c": "NC",
    "n.d": "ND", "ohio": "OH", "okla": "OK", "ore": "OR", "pa": "PA",
    "r.i": "RI", "s.c": "SC", "s.d": "SD", "tenn": "TN", "texas": "TX",
    "utah": "UT", "vt": "VT", "va": "VA", "wash": "WA", "w.va": "WV",
    "wis": "WI", "wyo": "WY",
}


def state_name_to_code(name):
    """Convert state name (possibly abbreviated, with footnote markers) to 2-letter code."""
    # Strip footnote markers like (a), (b,c), etc.
    clean = re.sub(r"\s*\([^)]*\)\s*", "", name).strip().rstrip(".")
    if clean.upper() in STATE_NAMES:
        return clean.upper()
    # Full name lookup
    code = NAME_TO_CODE.get(clean.lower())
    if code:
        return code
    # Abbreviated name lookup (e.g., "Ala" → AL, "Calif" → CA)
    code = STATE_ABBREVS.get(clean.lower())
    if code:
        return code
    return None


# ─── Phase A: TAXSIM35 (1977–2013) ──────────────────────────────────────────

TAXSIM_SSH_HOST = "taxsim35@taxsimssh.nber.org"
TAXSIM_HTTP_URL = "https://taxsim.nber.org/taxsim35/redirect.cgi"
TAXSIM_HTTP_BATCH_SIZE = 1800  # HTTP limited to ~2000 records


def run_taxsim_ssh(csv_input):
    """Submit CSV to TAXSIM35 via SSH (best for large batches)."""
    try:
        result = subprocess.run(
            [
                "ssh", "-T",
                "-o", "StrictHostKeyChecking=no",
                "-o", "UserKnownHostsFile=/dev/null",
                "-o", "ConnectTimeout=15",
                TAXSIM_SSH_HOST,
            ],
            input=csv_input,
            capture_output=True,
            text=True,
            timeout=600,
        )
        if result.returncode == 0 and result.stdout.strip():
            return result.stdout
        log.warning(f"SSH failed (rc={result.returncode}): {result.stderr[:300]}")
        return None
    except Exception as e:
        log.warning(f"SSH failed: {e}")
        return None


def run_taxsim_http_batch(csv_input):
    """Submit CSV to TAXSIM35 via HTTP in batches of ~1800 rows."""
    lines = csv_input.strip().split("\n")
    header = lines[0]
    data_lines = lines[1:]

    all_output_lines = []

    for batch_start in range(0, len(data_lines), TAXSIM_HTTP_BATCH_SIZE):
        batch = data_lines[batch_start : batch_start + TAXSIM_HTTP_BATCH_SIZE]
        batch_csv = header + "\n" + "\n".join(batch) + "\n"

        batch_num = batch_start // TAXSIM_HTTP_BATCH_SIZE + 1
        total_batches = (len(data_lines) + TAXSIM_HTTP_BATCH_SIZE - 1) // TAXSIM_HTTP_BATCH_SIZE
        if batch_num % 10 == 1:
            log.info(f"  HTTP batch {batch_num}/{total_batches}")

        try:
            # Multipart form upload
            boundary = "----TaxsimBoundary"
            body = (
                f"--{boundary}\r\n"
                f'Content-Disposition: form-data; name="txpydata.raw"; filename="txpydata.raw"\r\n'
                f"Content-Type: text/plain\r\n\r\n"
                f"{batch_csv}\r\n"
                f"--{boundary}--\r\n"
            ).encode("utf-8")

            req = urllib.request.Request(
                TAXSIM_HTTP_URL,
                data=body,
                headers={
                    "Content-Type": f"multipart/form-data; boundary={boundary}",
                    "User-Agent": "Mozilla/5.0 (research/academic)",
                },
            )
            with urllib.request.urlopen(req, timeout=60) as resp:
                result = resp.read().decode("utf-8")
                resp_lines = result.strip().split("\n")
                # Skip header on subsequent batches
                if batch_start == 0:
                    all_output_lines.extend(resp_lines)
                else:
                    # Skip first line if it's a header
                    start = 1 if resp_lines and not resp_lines[0][0].isdigit() else 0
                    all_output_lines.extend(resp_lines[start:])

        except Exception as e:
            log.warning(f"  HTTP batch {batch_num} failed: {e}")
            continue

    return "\n".join(all_output_lines) if all_output_lines else None


def run_taxsim_batch(csv_input):
    """Run TAXSIM35 — tries SSH first, falls back to HTTP batching."""
    log.info("Trying TAXSIM via SSH...")
    result = run_taxsim_ssh(csv_input)
    if result:
        log.info("SSH succeeded")
        return result

    log.info("SSH unavailable, falling back to HTTP batching...")
    result = run_taxsim_http_batch(csv_input)
    if result:
        log.info("HTTP batching succeeded")
        return result

    log.error("Both SSH and HTTP TAXSIM methods failed")
    return None


def detect_brackets_from_probes(probes_with_rates):
    """
    Detect bracket boundaries from marginal rate data.
    probes_with_rates: sorted list of (income, marginal_rate)
    Returns: list of [threshold, rate]
    """
    if not probes_with_rates:
        return []

    brackets = []
    prev_rate = None

    for income, rate in probes_with_rates:
        rate = round(rate, 6)
        if rate < 0:
            rate = 0.0  # Ignore negative marginal rates (refundable credits)
        if rate != prev_rate:
            if prev_rate is None:
                brackets.append([0, rate])
            else:
                brackets.append([income, rate])
            prev_rate = rate

    # Remove leading zero-rate bracket if followed by another starting at 0
    if len(brackets) > 1 and brackets[0][1] == 0.0 and brackets[1][0] == 0:
        brackets = brackets[1:]

    return brackets


def phase_a_taxsim(year_start=1977, year_end=2013):
    """Extract 1977–2013 brackets via TAXSIM35 marginal rate probing."""
    # TAXSIM mstat mapping
    mstat_map = {1: "Single", 2: "Married Filing Jointly"}

    # Generate all probe rows
    log.info(f"Phase A: Generating probes for {year_start}–{year_end}...")
    rows = []
    row_meta = []  # Track (state_code, year, mstat, income) per row
    row_id = 1

    for year in range(year_start, year_end + 1):
        for state_code, fips in STATE_FIPS.items():
            if state_code in NO_TAX_STATES:
                continue
            for mstat in [1, 2]:
                for income in INCOME_PROBES:
                    rows.append(f"{row_id},{year},{fips},{mstat},{income}")
                    row_meta.append((state_code, year, mstat, income))
                    row_id += 1

    log.info(f"Phase A: Generated {len(rows):,} probe rows, running TAXSIM35...")

    header = "taxsimid,year,state,mstat,pwages"
    csv_input = header + "\n" + "\n".join(rows) + "\n"

    output = run_taxsim_batch(csv_input)
    if not output:
        log.warning("Phase A: TAXSIM35 produced no output")
        return {}, {}

    # Parse output — standard output has 9 columns:
    # taxsimid, year, state, fiitax, siitax, fica, frate, srate, ficar
    log.info("Phase A: Parsing output...")
    probes_by_key = defaultdict(list)

    output_lines = output.strip().split("\n")
    start_idx = 0
    if output_lines and not output_lines[0][0].isdigit():
        start_idx = 1  # Skip header

    for line in output_lines[start_idx:]:
        try:
            parts = line.split(",")
            if len(parts) < 9:
                continue
            taxsimid = int(float(parts[0].strip()))
            srate = float(parts[7].strip()) / 100.0  # Convert percentage to decimal

            # Look up metadata
            idx = taxsimid - 1
            if idx < 0 or idx >= len(row_meta):
                continue
            state_code, year, mstat, income = row_meta[idx]

            probes_by_key[(state_code, year, mstat)].append((income, srate))
        except (ValueError, IndexError):
            continue

    # Detect brackets
    log.info("Phase A: Detecting brackets from marginal rates...")
    brackets_out = {}

    for (state_code, year, mstat), probes in probes_by_key.items():
        probes.sort(key=lambda x: x[0])
        brackets = detect_brackets_from_probes(probes)
        if not brackets:
            continue

        status = mstat_map[mstat]
        year_str = str(year)

        brackets_out.setdefault(state_code, {})
        brackets_out[state_code].setdefault(year_str, {})
        brackets_out[state_code][year_str][status] = brackets

    states_found = len(brackets_out)
    total_entries = sum(len(yrs) for yrs in brackets_out.values())
    log.info(f"Phase A: Extracted brackets for {states_found} states, {total_entries} state-years")
    return brackets_out, {}  # TAXSIM doesn't provide deduction data directly


# ─── Phase B: taxgraphs GitHub (2014–2024) ───────────────────────────────────

TAXGRAPHS_BASE = "https://raw.githubusercontent.com/hermantran/taxgraphs/master/data"


def fetch_taxgraphs_json(year, state_code):
    """Download and parse a state JSON from the taxgraphs repo."""
    url = f"{TAXGRAPHS_BASE}/{year}/state/{state_code}.json"
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=15) as resp:
            return json.loads(resp.read().decode())
    except Exception:
        return None


def normalize_taxgraphs_rate(rate_data):
    """
    Normalize taxgraphs rate data to {filing_status: [[threshold, rate], ...]}.
    Returns None for no-tax states.
    """
    if rate_data == 0 or rate_data == "none" or rate_data is None:
        return None

    if isinstance(rate_data, (int, float)):
        # Flat tax
        rate = float(rate_data)
        if rate <= 0:
            return None
        return {
            "Single": [[0, rate]],
            "Married Filing Jointly": [[0, rate]],
        }

    if isinstance(rate_data, list):
        # Bare list of [threshold, rate] — same brackets for all statuses
        brackets = [[float(t), float(r)] for t, r in rate_data]
        if not brackets or all(r == 0 for _, r in brackets):
            return None
        return {
            "Single": brackets,
            "Married Filing Jointly": list(brackets),
        }

    if isinstance(rate_data, dict):
        result = {}

        if "single" in rate_data:
            single = rate_data["single"]
            if isinstance(single, list):
                result["Single"] = [[float(t), float(r)] for t, r in single]
            elif isinstance(single, (int, float)) and float(single) > 0:
                result["Single"] = [[0, float(single)]]

        if "married" in rate_data:
            married = rate_data["married"]
            if isinstance(married, list):
                result["Married Filing Jointly"] = [[float(t), float(r)] for t, r in married]
            elif isinstance(married, (int, float)) and float(married) > 0:
                result["Married Filing Jointly"] = [[0, float(married)]]

        # Copy if only one status available
        if "Single" in result and "Married Filing Jointly" not in result:
            result["Married Filing Jointly"] = list(result["Single"])
        elif "Married Filing Jointly" in result and "Single" not in result:
            result["Single"] = list(result["Married Filing Jointly"])

        return result if result else None

    return None


def phase_b_taxgraphs(year_start=2014, year_end=2024):
    """Extract brackets from taxgraphs GitHub repo for 2014–2024."""
    brackets_out = {}
    deductions_out = {}

    total = (year_end - year_start + 1) * len(STATE_NAMES)
    done = 0

    for year in range(year_start, year_end + 1):
        for state_code in STATE_NAMES:
            done += 1
            if done % 100 == 0:
                log.info(f"Phase B: {done}/{total} ({state_code} {year})")

            data = fetch_taxgraphs_json(year, state_code)
            if not data:
                continue

            try:
                income = data.get("taxes", {}).get("income", {})
                rate_data = income.get("rate", 0)
                brackets = normalize_taxgraphs_rate(rate_data)

                year_str = str(year)
                brackets_out.setdefault(state_code, {})

                if brackets:
                    brackets_out[state_code][year_str] = brackets

                # Standard deductions
                deductions = income.get("deductions", {})
                if not isinstance(deductions, dict):
                    deductions = {}
                std_ded_node = deductions.get("standardDeduction", {})
                if not isinstance(std_ded_node, dict):
                    std_ded_node = {}
                amount = std_ded_node.get("amount", {})

                if amount:
                    ded_entry = {}
                    if isinstance(amount, (int, float)):
                        # Scalar: same deduction for all filing statuses
                        ded_entry["Single"] = float(amount)
                        ded_entry["Married Filing Jointly"] = float(amount)
                    elif isinstance(amount, dict):
                        for src_key, dst_key in [("single", "Single"), ("married", "Married Filing Jointly")]:
                            val = amount.get(src_key)
                            if val is not None and str(val).lower() not in ("n.a.", "n/a", "none", ""):
                                try:
                                    ded_entry[dst_key] = float(val)
                                except (ValueError, TypeError):
                                    pass

                    if ded_entry:
                        deductions_out.setdefault(state_code, {})
                        deductions_out[state_code][year_str] = ded_entry

            except Exception as e:
                log.warning(f"Phase B: Error parsing {state_code} {year}: {e}")

    log.info(f"Phase B: Extracted data for {len(brackets_out)} states")
    return brackets_out, deductions_out


# ─── Phase C: Tax Foundation Excel (2023–2026) ──────────────────────────────

TAX_FOUNDATION_EXCEL_URL = (
    "https://taxfoundation.org/wp-content/uploads/2026/02/"
    "2026-State-Individual-Income-Tax-Rates-Brackets.xlsx"
)


def download_tax_foundation_excel():
    """Download the Tax Foundation Excel file."""
    os.makedirs(CACHE_DIR, exist_ok=True)
    excel_path = os.path.join(CACHE_DIR, "state_tax_brackets.xlsx")
    if os.path.exists(excel_path):
        log.info(f"Tax Foundation Excel already cached: {excel_path}")
        return excel_path

    log.info(f"Downloading Tax Foundation Excel...")
    try:
        req = urllib.request.Request(
            TAX_FOUNDATION_EXCEL_URL,
            headers={"User-Agent": "Mozilla/5.0 (research/academic)"},
        )
        with urllib.request.urlopen(req, timeout=30) as resp:
            with open(excel_path, "wb") as f:
                f.write(resp.read())
        log.info(f"Downloaded to {excel_path}")
        return excel_path
    except Exception as e:
        log.error(f"Failed to download Tax Foundation Excel: {e}")
        return None


def _parse_number(val):
    """Try to parse a numeric value from a cell, return None on failure."""
    if val is None:
        return None
    s = str(val).strip()
    if s.lower() in ("", "n.a.", "n/a", "-", "none", "…", "–"):
        return None
    # Strip $, %, commas
    s = s.replace("$", "").replace("%", "").replace(",", "").strip()
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def parse_excel_sheet(ws, year):
    """
    Parse a Tax Foundation Excel sheet.
    Column layout (verified):
      0: State  1: Single Rate  2: ">"  3: Single Bracket
      4: Married Rate  5: ">"  6: Married Bracket
      7: Std Ded (Single)  8: Std Ded (Couple)
      9: Personal Exempt (Single)  10: (Couple)  11: (Dependent)
    """
    brackets = {}
    deductions = {}

    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return brackets, deductions

    # Find header row (Row 2 has "State")
    header_idx = None
    for i, row in enumerate(all_rows):
        if row[0] and str(row[0]).strip().lower() == "state":
            header_idx = i
            break

    if header_idx is None:
        log.warning(f"Could not find header row for year {year}")
        return brackets, deductions

    current_code = None
    first_row_for_state = True

    for row in all_rows[header_idx + 1 :]:
        if not row or len(row) < 7:
            continue

        # Check for state name in column 0
        state_cell = str(row[0] or "").strip()

        # Skip empty rows, footnote-only rows like "(a, b, c)", and "\xa0"
        if state_cell and state_cell != "\xa0":
            # Footnote rows start with "(" — they belong to the current state
            if not state_cell.startswith("("):
                code = state_name_to_code(state_cell)
                if code:
                    current_code = code
                    first_row_for_state = True
                    brackets.setdefault(code, {"Single": [], "Married Filing Jointly": []})

        if not current_code:
            continue

        # Extract rates and brackets from fixed column positions
        single_rate = _parse_number(row[1])
        single_bracket = _parse_number(row[3])
        married_rate = _parse_number(row[4])
        married_bracket = _parse_number(row[6])

        # Skip rows with "none" or no rate data
        rate_str = str(row[1] or "").strip().lower()
        if rate_str in ("none", "", "\xa0"):
            # No-tax state or empty row
            if first_row_for_state and rate_str == "none":
                first_row_for_state = False
            continue

        if single_rate is None:
            continue

        # Rates are already decimal in the Excel (e.g., 0.02 for 2%)
        # but some might be expressed as 2.0 (meaning 2%) — normalize
        if single_rate > 1:
            single_rate /= 100.0
        threshold_s = single_bracket if single_bracket is not None else 0.0
        brackets[current_code]["Single"].append([threshold_s, single_rate])

        if married_rate is not None:
            if married_rate > 1:
                married_rate /= 100.0
            threshold_m = married_bracket if married_bracket is not None else 0.0
            brackets[current_code]["Married Filing Jointly"].append([threshold_m, married_rate])
        else:
            # Same rate for married (flat-tax or same brackets)
            brackets[current_code]["Married Filing Jointly"].append([threshold_s, single_rate])

        # Standard deduction (first data row of state only)
        if first_row_for_state:
            std_single = _parse_number(row[7]) if len(row) > 7 else None
            std_couple = _parse_number(row[8]) if len(row) > 8 else None
            if std_single is not None:
                deductions[current_code] = {"Single": std_single}
                if std_couple is not None:
                    deductions[current_code]["Married Filing Jointly"] = std_couple
                else:
                    deductions[current_code]["Married Filing Jointly"] = std_single * 2
            first_row_for_state = False

    # Sort brackets by threshold
    for code in brackets:
        for status in brackets[code]:
            brackets[code][status].sort(key=lambda x: x[0])

    return brackets, deductions


def phase_c_excel():
    """Extract brackets from Tax Foundation Excel (2023–2026)."""
    try:
        import openpyxl
    except ImportError:
        log.error("Phase C requires openpyxl: pip install openpyxl")
        return {}, {}

    excel_path = download_tax_foundation_excel()
    if not excel_path:
        log.warning("Skipping Phase C: Excel not available")
        return {}, {}

    brackets_out = {}
    deductions_out = {}

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)

        for sheet_name in wb.sheetnames:
            # Extract year from sheet name
            match = re.search(r"20\d{2}", sheet_name)
            if not match:
                continue
            year = int(match.group())
            year_str = str(year)

            log.info(f"Phase C: Parsing sheet '{sheet_name}' (year {year})")
            ws = wb[sheet_name]
            sheet_brackets, sheet_deductions = parse_excel_sheet(ws, year)

            for code, statuses in sheet_brackets.items():
                if any(statuses[s] for s in statuses):
                    brackets_out.setdefault(code, {})
                    brackets_out[code][year_str] = statuses

            for code, deds in sheet_deductions.items():
                deductions_out.setdefault(code, {})
                deductions_out[code][year_str] = deds

        wb.close()
    except Exception as e:
        log.error(f"Phase C: Error processing Excel: {e}")

    log.info(f"Phase C: Extracted data for {len(brackets_out)} states")
    return brackets_out, deductions_out


# ─── Phase D: Cross-validation ──────────────────────────────────────────────


def cross_validate(source_a, source_b, overlap_years, label_a, label_b):
    """Compare brackets between two sources at overlapping years."""
    mismatches = 0
    matches = 0

    for state_code in STATE_NAMES:
        for year in overlap_years:
            year_str = str(year)
            a_data = source_a.get(state_code, {}).get(year_str, {})
            b_data = source_b.get(state_code, {}).get(year_str, {})

            if not a_data or not b_data:
                continue

            for status in ["Single", "Married Filing Jointly"]:
                a_brackets = a_data.get(status, [])
                b_brackets = b_data.get(status, [])

                if len(a_brackets) != len(b_brackets):
                    log.debug(
                        f"  {state_code} {year} {status}: "
                        f"{label_a} has {len(a_brackets)} brackets, "
                        f"{label_b} has {len(b_brackets)}"
                    )
                    mismatches += 1
                    continue

                for i, (a, b) in enumerate(zip(a_brackets, b_brackets)):
                    if abs(a[1] - b[1]) > 0.002:
                        log.debug(
                            f"  {state_code} {year} {status} bracket {i}: "
                            f"rate {a[1]:.4f} vs {b[1]:.4f}"
                        )
                        mismatches += 1
                    else:
                        matches += 1

    log.info(
        f"Cross-validation ({label_a} vs {label_b}): "
        f"{matches} matches, {mismatches} mismatches"
    )


# ─── Phase E: Merge and Output ──────────────────────────────────────────────


def merge_and_output(bracket_sources, deduction_sources, output_path):
    """Merge all sources into final JSON and write to disk."""
    final = {}

    for state_code, name in sorted(STATE_NAMES.items()):
        entry = {
            "name": name,
            "has_income_tax": state_code not in NO_TAX_STATES,
            "lt_exclusion_pct": LT_EXCLUSION_PCT.get(state_code, 0.0),
            "brackets": {},
            "standard_deductions": {},
        }

        # Merge brackets (later sources override earlier for same year)
        for source in bracket_sources:
            state_data = source.get(state_code, {})
            for year_str, year_brackets in state_data.items():
                entry["brackets"][year_str] = year_brackets

        # Merge deductions
        for ded_source in deduction_sources:
            state_deds = ded_source.get(state_code, {})
            for year_str, deds in state_deds.items():
                entry["standard_deductions"][year_str] = deds

        final[state_code] = entry

    # Write
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, "w") as f:
        json.dump(final, f, indent=2, sort_keys=False)

    size_kb = os.path.getsize(output_path) / 1024
    total_years = sum(len(v["brackets"]) for v in final.values())
    states_with_data = sum(1 for v in final.values() if v["brackets"])
    log.info(f"Wrote {output_path} ({size_kb:.1f} KB)")
    log.info(f"States with bracket data: {states_with_data}/{len(STATE_NAMES)}")
    log.info(f"Total state-year entries: {total_years}")


# ─── Main ───────────────────────────────────────────────────────────────────


def main():
    parser = argparse.ArgumentParser(description="Build state income tax bracket data")
    parser.add_argument(
        "--skip-taxsim",
        action="store_true",
        help="Skip TAXSIM35 extraction for 1977–2013",
    )
    parser.add_argument(
        "--skip-taxgraphs", action="store_true", help="Skip taxgraphs GitHub extraction"
    )
    parser.add_argument(
        "--skip-excel", action="store_true", help="Skip Tax Foundation Excel extraction"
    )
    parser.add_argument("--output", default=OUTPUT_PATH, help=f"Output path (default: {OUTPUT_PATH})")
    args = parser.parse_args()

    all_bracket_sources = []
    all_deduction_sources = []

    # Phase A: TAXSIM (1977–2013)
    if not args.skip_taxsim:
        log.info("=" * 60)
        log.info("Phase A: TAXSIM35 (1977–2013)")
        log.info("=" * 60)
        b, d = phase_a_taxsim()
        all_bracket_sources.append(b)
        all_deduction_sources.append(d)
    else:
        log.info("Phase A: Skipped (use --skip-taxsim to skip)")
        all_bracket_sources.append({})
        all_deduction_sources.append({})

    # Phase B: taxgraphs (2014–2024)
    if not args.skip_taxgraphs:
        log.info("=" * 60)
        log.info("Phase B: taxgraphs GitHub (2014–2024)")
        log.info("=" * 60)
        b, d = phase_b_taxgraphs()
        all_bracket_sources.append(b)
        all_deduction_sources.append(d)
    else:
        log.info("Phase B: Skipped")
        all_bracket_sources.append({})
        all_deduction_sources.append({})

    # Phase C: Tax Foundation Excel (2023–2026)
    if not args.skip_excel:
        log.info("=" * 60)
        log.info("Phase C: Tax Foundation Excel (2023–2026)")
        log.info("=" * 60)
        b, d = phase_c_excel()
        all_bracket_sources.append(b)
        all_deduction_sources.append(d)
    else:
        log.info("Phase C: Skipped")
        all_bracket_sources.append({})
        all_deduction_sources.append({})

    # Phase D: Cross-validation
    log.info("=" * 60)
    log.info("Phase D: Cross-validation")
    log.info("=" * 60)

    src_b, src_c = all_bracket_sources[1], all_bracket_sources[2]
    if src_b and src_c:
        cross_validate(src_b, src_c, range(2023, 2025), "taxgraphs", "Excel")
    else:
        log.info("Insufficient sources for cross-validation")

    # Phase E: Merge and Output
    log.info("=" * 60)
    log.info("Phase E: Merge and Output")
    log.info("=" * 60)

    merge_and_output(all_bracket_sources, all_deduction_sources, args.output)
    log.info("Done!")


if __name__ == "__main__":
    main()
